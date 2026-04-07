import json
import logging
import os
import secrets
import sqlite3
import sys
from threading import Lock
from logging.handlers import RotatingFileHandler
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO
from pathlib import Path
from typing import Any

from flask import Flask, flash, g, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from werkzeug.exceptions import HTTPException
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from crm_constants import (
    AMOUNT_CONFIDENCE_LABELS,
    APPROVAL_STATUS_LABELS,
    CONTRACT_DEVIATION_THRESHOLD_PCT,
    CONTRACT_STATUS_LABELS,
    CUSTOMER_FOLLOW_METHOD_LABELS,
    CUSTOMER_STALE_FOLLOW_DAYS,
    CUSTOMER_STATUS_LABELS,
    CUSTOMER_TIER_LABELS,
    HEALTH_LABELS,
    INVOICE_STATUS_LABELS,
    LOST_REASON_LABELS,
    MILESTONE_STATUS_IMPORT_MAP,
    MILESTONE_STATUS_LABELS,
    OPPORTUNITY_STAGE_LABELS,
    OPPORTUNITY_STATUS_LABELS,
    PM_ROLE_CODE,
    PRIORITY_IMPORT_MAP,
    PRIORITY_LABELS,
    PROJECT_RECYCLE_DAYS,
    PROJECT_PHASE_STATUS_LABELS,
    PROJECT_STAGE_LABELS,
    PROJECT_STATUS_LABELS,
    PROJECT_TYPE_LABELS,
    RECEIVABLE_STATUS_LABELS,
    RISK_LEVEL_LABELS,
    RISK_STATUS_LABELS,
    ROLE_LABELS,
    ROLE_PERMISSIONS,
    TASK_STATUS_IMPORT_MAP,
    TASK_STATUS_LABELS,
    VALID_PERMISSION_KEYS,
    permission_catalog_by_group,
)
from crm_utils import (
    env_value,
    format_file_size,
    in_clause_params,
    like_kw,
    load_env_file,
    max_upload_bytes,
    normalize_stage,
    now_iso,
    parse_date_text,
    resolve_app_port,
    slug_role_code,
)

try:
    import psycopg
    from psycopg.rows import dict_row
except ImportError:  # pragma: no cover
    psycopg = None
    dict_row = None

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None
    load_workbook = None


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "crm.db"
UPLOAD_DIR = BASE_DIR / "uploads"
LOG_DIR = BASE_DIR / "logs"

load_env_file(BASE_DIR / ".env")

APP_HOST = os.getenv("APP_HOST", "0.0.0.0").strip() or "0.0.0.0"
APP_PORT = resolve_app_port()
DB_BACKEND = "postgres"
PG_HOST = env_value("PG_HOST", "PGHOST", default="192.168.0.103")
PG_PORT = int(env_value("PG_PORT", "PGPORT", default="23083"))
PG_DATABASE = env_value("PG_DATABASE", "PGDATABASE", default="zqq_test")
PG_USER = env_value("PG_USERNAME", "PGUSER", default="zqq")
PG_PASSWORD = env_value("PG_PASSWORD", "PGPASSWORD", default="qq199499")
DEFAULT_ADMIN_USER = os.getenv("DEFAULT_ADMIN_USER", "admin")
DEFAULT_ADMIN_PASSWORD = os.getenv("DEFAULT_ADMIN_PASSWORD", "admin123")

# 登录验证码：字母数字（排除易混淆的 0/O、1/I/L）
_CAPTCHA_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"


def prepare_login_captcha() -> dict[str, Any]:
    code = "".join(secrets.choice(_CAPTCHA_ALPHABET) for _ in range(4))
    session["captcha_code"] = code
    return {"captcha_display": code}


def _role_is_active(r: dict[str, Any]) -> bool:
    return r.get("is_active") not in (False, 0, "0")


def _normalize_perm_keys(keys: list[str]) -> list[str]:
    cleaned = [k for k in keys if k in VALID_PERMISSION_KEYS]
    if "*" in cleaned:
        return ["*"]
    return cleaned


def _replace_role_permissions(role_id: int, perm_keys: list[str]) -> None:
    execute("DELETE FROM role_permissions WHERE role_id = ?", (role_id,))
    for pk in perm_keys:
        execute("INSERT INTO role_permissions(role_id, perm_key) VALUES (?, ?)", (role_id, pk))


def _clean_role_ids(role_ids: list[Any]) -> list[int]:
    cleaned: list[int] = []
    seen: set[int] = set()
    for raw in role_ids:
        try:
            rid = int(raw)
        except (TypeError, ValueError):
            continue
        if rid <= 0 or rid in seen:
            continue
        seen.add(rid)
        cleaned.append(rid)
    return cleaned


def fetch_active_roles_by_ids(role_ids: list[Any]) -> list[dict[str, Any]]:
    cleaned = _clean_role_ids(role_ids)
    if not cleaned:
        return []
    placeholders = ", ".join("?" for _ in cleaned)
    rows = fetchall(
        f"SELECT id, code, name, sort_order, is_active FROM roles WHERE id IN ({placeholders})",
        tuple(cleaned),
    )
    rows = [row for row in rows if _role_is_active(row)]
    rows.sort(key=lambda row: (int(row.get("sort_order") or 0), str(row.get("name") or ""), int(row.get("id") or 0)))
    return rows


def get_user_role_rows(user_id: int) -> list[dict[str, Any]]:
    cached_map = getattr(g, "_user_role_rows_cache", None)
    if cached_map is None:
        cached_map = {}
        g._user_role_rows_cache = cached_map
    if user_id in cached_map:
        return cached_map[user_id]
    rows = fetchall(
        """
        SELECT r.id, r.code, r.name, r.sort_order, r.is_active
        FROM user_roles ur
        JOIN roles r ON r.id = ur.role_id
        WHERE ur.user_id = ?
        ORDER BY r.sort_order, r.name, r.id
        """,
        (user_id,),
    )
    rows = [row for row in rows if _role_is_active(row)]
    if rows:
        cached_map[user_id] = rows
        return rows
    fallback = fetchone(
        """
        SELECT r.id AS id, r.code AS code, r.name AS name, r.sort_order AS sort_order, r.is_active AS is_active
        FROM users u
        LEFT JOIN roles r ON r.id = u.role_id
        WHERE u.id = ?
        """,
        (user_id,),
    )
    if fallback and fallback.get("id") is not None:
        cached_map[user_id] = [fallback]
        return cached_map[user_id]
    if fallback and fallback.get("code"):
        cached_map[user_id] = [fallback]
        return cached_map[user_id]
    user = fetchone("SELECT role FROM users WHERE id = ?", (user_id,))
    if user and user.get("role"):
        role_row = fetchone(
            "SELECT id, code, name, sort_order, is_active FROM roles WHERE code = ?",
            (str(user.get("role") or "").strip(),),
        )
        if role_row and _role_is_active(role_row):
            cached_map[user_id] = [role_row]
            return cached_map[user_id]
    cached_map[user_id] = []
    return cached_map[user_id]


def get_user_role_labels(user_id: int) -> list[str]:
    return [str(row.get("name") or row.get("code") or "").strip() for row in get_user_role_rows(user_id) if str(row.get("name") or row.get("code") or "").strip()]


def get_user_role_codes(user_id: int) -> list[str]:
    return [str(row.get("code") or "").strip() for row in get_user_role_rows(user_id) if str(row.get("code") or "").strip()]


def set_user_roles(user_id: int, role_ids: list[Any]) -> list[dict[str, Any]]:
    roles = fetch_active_roles_by_ids(role_ids)
    if not roles:
        raise ValueError("至少需要一个有效角色。")
    execute("DELETE FROM user_roles WHERE user_id = ?", (user_id,))
    for role in roles:
        execute("INSERT INTO user_roles(user_id, role_id) VALUES (?, ?)", (user_id, int(role["id"])))
    primary = roles[0]
    execute(
        "UPDATE users SET role = ?, role_id = ?, updated_at = ? WHERE id = ?",
        (str(primary["code"]).strip(), int(primary["id"]), now_iso(), user_id),
    )
    return roles


app = Flask(__name__)
app.config["SECRET_KEY"] = "simple-crm-secret-key"
app.config["UPLOAD_FOLDER"] = str(UPLOAD_DIR)
app.config["MAX_CONTENT_LENGTH"] = max_upload_bytes()
# 登录态采用“最近有效操作”过期：只有真正的业务请求才会续期。
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)
app.config["SESSION_REFRESH_EACH_REQUEST"] = False
# 路由模块统一从 `app` 导入共享对象；直接执行 `python app.py` 时这里提供同名别名，
# 避免解释器把主模块当成 `__main__` 再重复导入一份 `app.py`。
sys.modules.setdefault("app", sys.modules[__name__])


def setup_file_logging(application: Flask) -> None:
    """将应用日志写入 logs 目录，便于排查错误（含异常堆栈）。"""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOG_DIR / "crm.log"
    file_handler = RotatingFileHandler(
        str(log_path),
        maxBytes=10 * 1024 * 1024,
        backupCount=10,
        encoding="utf-8",
    )
    level_name = (os.getenv("CRM_LOG_LEVEL") or "INFO").strip().upper()
    level = getattr(logging, level_name, logging.INFO)
    file_handler.setLevel(level)
    file_handler.setFormatter(
        logging.Formatter(
            "%(asctime)s [%(levelname)s] %(name)s %(filename)s:%(lineno)d — %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
    )
    application.logger.setLevel(level)
    if not getattr(application, "_crm_file_log_handler", None):
        application.logger.addHandler(file_handler)
        application._crm_file_log_handler = file_handler
    application.logger.propagate = False


setup_file_logging(app)


SESSION_IDLE_TIMEOUT_SECONDS = int(timedelta(minutes=30).total_seconds())


def _session_last_activity_ts() -> float | None:
    raw = session.get("last_activity_ts")
    if raw is None:
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _touch_session_activity() -> None:
    session["last_activity_ts"] = datetime.utcnow().timestamp()
    session.permanent = True
    session.modified = True


@app.errorhandler(Exception)
def handle_unhandled_exception(e: Exception):
    """HTTP 异常（如 404）正常返回；其它异常写入 logs 并返回 500。"""
    if isinstance(e, HTTPException):
        return e
    app.logger.exception("未处理异常: %s", e)
    return (
        "<!DOCTYPE html><html lang=\"zh-CN\"><head><meta charset=\"utf-8\"><title>错误</title></head>"
        "<body style=\"font-family:system-ui,sans-serif;padding:2rem;\">"
        "<h1>服务器错误</h1><p>系统已记录异常，请稍后再试或联系管理员。</p>"
        "<p style=\"color:#64748b;font-size:14px;\">查看日志文件：logs/crm.log</p>"
        "</body></html>",
        500,
    )


@app.errorhandler(413)
def request_entity_too_large(_e: Any):
    flash(
        "上传的文件超过当前允许的大小。请压缩后重试；若需更大上限，可在服务器环境变量中设置 MAX_UPLOAD_MB（单位 MB，默认 200，最大 2048）。",
        "warning",
    )
    ref = request.referrer
    if ref:
        return redirect(ref)
    return redirect(url_for("dashboard"))


@app.context_processor
def inject_labels() -> dict[str, Any]:
    current_role_rows = get_user_role_rows(int(session["user_id"])) if session.get("user_id") else []
    current_role_labels = [str(row.get("name") or row.get("code") or "").strip() for row in current_role_rows if str(row.get("name") or row.get("code") or "").strip()]
    current_role_codes = [str(row.get("code") or "").strip() for row in current_role_rows if str(row.get("code") or "").strip()]
    primary_role_code = current_role_codes[0] if current_role_codes else session.get("role", "normal")
    primary_role_label = current_role_labels[0] if current_role_labels else ROLE_LABELS.get(primary_role_code, primary_role_code)
    return {
        "project_type_labels": PROJECT_TYPE_LABELS,
        "project_status_labels": PROJECT_STATUS_LABELS,
        "task_status_labels": TASK_STATUS_LABELS,
        "priority_labels": PRIORITY_LABELS,
        "milestone_status_labels": MILESTONE_STATUS_LABELS,
        "risk_level_labels": RISK_LEVEL_LABELS,
        "risk_status_labels": RISK_STATUS_LABELS,
        "health_labels": HEALTH_LABELS,
        "project_stage_labels": PROJECT_STAGE_LABELS,
        "project_phase_status_labels": PROJECT_PHASE_STATUS_LABELS,
        "opportunity_status_labels": OPPORTUNITY_STATUS_LABELS,
        "opportunity_stage_labels": OPPORTUNITY_STAGE_LABELS,
        "amount_confidence_labels": AMOUNT_CONFIDENCE_LABELS,
        "lost_reason_labels": LOST_REASON_LABELS,
        "contract_status_labels": CONTRACT_STATUS_LABELS,
        "contract_deviation_threshold_pct": CONTRACT_DEVIATION_THRESHOLD_PCT,
        "invoice_status_labels": INVOICE_STATUS_LABELS,
        "approval_status_labels": APPROVAL_STATUS_LABELS,
        "receivable_status_labels": RECEIVABLE_STATUS_LABELS,
        "customer_tier_labels": CUSTOMER_TIER_LABELS,
        "customer_status_labels": CUSTOMER_STATUS_LABELS,
        "customer_follow_method_labels": CUSTOMER_FOLLOW_METHOD_LABELS,
        "role_labels": ROLE_LABELS,
        "current_user": session.get("display_name") or session.get("username"),
        "current_role": primary_role_code,
        "current_role_label": " / ".join(current_role_labels) if current_role_labels else primary_role_label,
        "current_role_labels": current_role_labels,
        "current_role_codes": current_role_codes,
        "current_user_identity_values": sorted(current_user_identity_values()),
        "is_admin": session_is_system_admin(),
        "default_admin_username": DEFAULT_ADMIN_USER,
        "can": has_module_permission,
    }


def uses_postgres() -> bool:
    return DB_BACKEND == "postgres"


def create_connection():
    if uses_postgres():
        if psycopg is None:
            raise RuntimeError("当前配置为 PostgreSQL，但缺少 psycopg 依赖。请安装依赖后再启动。")
        return psycopg.connect(
            host=PG_HOST,
            port=PG_PORT,
            dbname=PG_DATABASE,
            user=PG_USER,
            password=PG_PASSWORD,
            row_factory=dict_row,
        )
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


_DB_BOOTSTRAP_LOCK = Lock()
_DB_BOOTSTRAPPED_KEYS: set[tuple[str, ...]] = set()
_ATTACHMENT_FILE_SIZE_LOCK = Lock()
_ATTACHMENT_FILE_SIZE_CACHE: dict[str, int] = {}


def _db_bootstrap_key() -> tuple[str, ...]:
    if uses_postgres():
        return (
            "postgres",
            PG_HOST,
            str(PG_PORT),
            PG_DATABASE,
            PG_USER,
            PG_PASSWORD,
        )
    if DB_PATH.exists():
        stat = DB_PATH.stat()
        return ("sqlite", str(DB_PATH.resolve()), str(stat.st_mtime_ns), str(stat.st_size))
    return ("sqlite", str(DB_PATH.resolve()), "missing")


def _bootstrap_connection_schema(conn: Any) -> None:
    key = _db_bootstrap_key()
    if key in _DB_BOOTSTRAPPED_KEYS:
        return
    with _DB_BOOTSTRAP_LOCK:
        if key in _DB_BOOTSTRAPPED_KEYS:
            return
        _ensure_roles_schema(conn)
        _ensure_project_progress_schema(conn)
        _ensure_project_members_schema(conn)
        _ensure_project_deleted_at_schema(conn)
        _ensure_approval_notes_schema(conn)
        _ensure_customer_engagement_schema(conn)
        _ensure_opportunity_pipeline_schema(conn)
        _ensure_contracts_extended_schema(conn)
        _ensure_project_phases_schema(conn)
        _ensure_query_indexes(conn)
        _DB_BOOTSTRAPPED_KEYS.add(key)


def q(sql_text: str) -> str:
    if uses_postgres():
        return sql_text.replace("?", "%s")
    return sql_text


def sql_bool(value: bool) -> Any:
    """按当前数据库后端生成布尔值参数。"""
    return value if uses_postgres() else (1 if value else 0)


def sql_user_order_clause(alias: str = "u") -> str:
    """用户名/姓名排序规则，PostgreSQL 与 SQLite 保持一致的结果。"""
    return f"{alias}.display_name" if uses_postgres() else f"{alias}.display_name COLLATE NOCASE"


def as_dict(row: Any) -> dict[str, Any]:
    if row is None:
        return {}
    if isinstance(row, dict):
        return row
    return dict(row)


def fetchone(sql_text: str, params: tuple[Any, ...] = ()) -> dict[str, Any] | None:
    cursor = get_db().cursor()
    cursor.execute(q(sql_text), params)
    row = cursor.fetchone()
    return as_dict(row) if row else None


def fetchall(sql_text: str, params: tuple[Any, ...] = ()) -> list[dict[str, Any]]:
    cursor = get_db().cursor()
    cursor.execute(q(sql_text), params)
    return [as_dict(row) for row in cursor.fetchall()]


def execute(sql_text: str, params: tuple[Any, ...] = ()) -> None:
    cursor = get_db().cursor()
    cursor.execute(q(sql_text), params)


def execute_returning_id(sql_text: str, params: tuple[Any, ...]) -> int:
    cursor = get_db().cursor()
    cursor.execute(q(sql_text), params)
    row = cursor.fetchone()
    if not row:
        raise RuntimeError("创建记录失败，未返回ID。")
    row_data = as_dict(row)
    return int(row_data["id"])


def _ensure_roles_schema(conn: Any) -> None:
    """创建 roles / role_permissions，补全 users.role_id，并在首次运行时写入内置角色与权限。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS roles (
                id SERIAL PRIMARY KEY,
                code VARCHAR(40) NOT NULL UNIQUE,
                name VARCHAR(120) NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                is_system BOOLEAN NOT NULL DEFAULT FALSE,
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                updated_at TIMESTAMP NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS role_permissions (
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                perm_key VARCHAR(80) NOT NULL,
                PRIMARY KEY (role_id, perm_key)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS user_roles (
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                PRIMARY KEY (user_id, role_id)
            )
            """
        )
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS role_id INTEGER REFERENCES roles(id)")
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS roles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                is_system INTEGER NOT NULL DEFAULT 0,
                is_active INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS role_permissions (
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                perm_key TEXT NOT NULL,
                PRIMARY KEY (role_id, perm_key)
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS user_roles (
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                role_id INTEGER NOT NULL REFERENCES roles(id) ON DELETE CASCADE,
                PRIMARY KEY (user_id, role_id)
            )
            """
        )
        cur.execute("PRAGMA table_info(users)")
        user_cols = [row[1] for row in cur.fetchall()]
        if "role_id" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN role_id INTEGER REFERENCES roles(id)")

    cur.execute(q("SELECT COUNT(1) AS c FROM roles"), ())
    row = cur.fetchone()
    n_roles = int(as_dict(row)["c"]) if row else 0
    cur.execute(q("SELECT COALESCE(MAX(sort_order), 0) AS m FROM roles"), ())
    row = cur.fetchone()
    next_sort_order = int(as_dict(row)["m"]) if row else 0

    for code, perms in ROLE_PERMISSIONS.items():
        name = ROLE_LABELS.get(code, code)
        cur.execute(q("SELECT id FROM roles WHERE code = ?"), (code,))
        rid_row = cur.fetchone()
        if rid_row is None:
            if n_roles == 0:
                sort_order = next_sort_order
                next_sort_order += 1
            else:
                next_sort_order += 1
                sort_order = next_sort_order
            cur.execute(
                q(
                    """
                    INSERT INTO roles(code, name, sort_order, is_system, is_active, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """
                ),
                (code, name, sort_order, sql_bool(True), sql_bool(True), now_iso()),
            )
            cur.execute(q("SELECT id FROM roles WHERE code = ?"), (code,))
            rid_row = cur.fetchone()
        elif rid_row and n_roles > 0:
            rid = int(as_dict(rid_row)["id"])
            cur.execute(
                q("UPDATE roles SET name = ?, is_system = ?, is_active = ? WHERE id = ?"),
                (name, sql_bool(True), sql_bool(True), rid),
            )
        if not rid_row:
            continue
        rid = int(as_dict(rid_row)["id"])
        cur.execute(q("SELECT perm_key FROM role_permissions WHERE role_id = ?"), (rid,))
        existing_perm_rows = cur.fetchall() or []
        existing_perms = {str(as_dict(p)["perm_key"]) for p in existing_perm_rows}
        for pk in perms:
            if pk in existing_perms:
                continue
            cur.execute(
                q("INSERT INTO role_permissions(role_id, perm_key) VALUES (?, ?)"),
                (rid, pk),
            )

    # 同步用户 role_id（旧数据仅有 role 字符串）
    cur.execute(
        q(
            """
            UPDATE users
            SET role_id = (SELECT id FROM roles WHERE roles.code = users.role LIMIT 1)
            WHERE role_id IS NULL AND role IS NOT NULL AND role != ''
            """
        ),
        (),
    )
    if uses_postgres():
        cur.execute(
            """
            INSERT INTO user_roles(user_id, role_id)
            SELECT id, role_id
            FROM users
            WHERE role_id IS NOT NULL
            ON CONFLICT DO NOTHING
            """
        )
    else:
        cur.execute(
            """
            INSERT OR IGNORE INTO user_roles(user_id, role_id)
            SELECT id, role_id
            FROM users
            WHERE role_id IS NOT NULL
            """
        )
    # 历史「超级管理员」与「管理员」合并为同一角色，避免重复
    cur.execute(q("SELECT id FROM roles WHERE code = ?"), ("super_admin",))
    row_sa = cur.fetchone()
    if row_sa:
        sa_id = int(as_dict(row_sa)["id"])
        cur.execute(q("SELECT id FROM roles WHERE code = ?"), ("admin",))
        row_ad = cur.fetchone()
        if row_ad:
            admin_id = int(as_dict(row_ad)["id"])
            cur.execute(
                q(
                    """
                    UPDATE users SET role = 'admin', role_id = ?
                    WHERE role = 'super_admin' OR role_id = ?
                    """
                ),
                (admin_id, sa_id),
            )
            cur.execute(q("DELETE FROM role_permissions WHERE role_id = ?"), (sa_id,))
            cur.execute(q("DELETE FROM roles WHERE id = ?"), (sa_id,))
    conn.commit()


def _ensure_project_progress_schema(conn: Any) -> None:
    """项目进展记录表：项目经理填写，仅管理员可删。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_progress_entries (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                body TEXT NOT NULL,
                created_by INTEGER REFERENCES users(id),
                created_at TIMESTAMP NOT NULL
            )
            """
        )
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_progress_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                body TEXT NOT NULL,
                created_by INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
                FOREIGN KEY(created_by) REFERENCES users(id)
            )
            """
        )
    conn.commit()


def _ensure_project_members_schema(conn: Any) -> None:
    """项目成员：项目经理可将任意用户加入项目。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_members (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                created_at TIMESTAMP NOT NULL,
                UNIQUE(project_id, user_id)
            )
            """
        )
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(project_id, user_id),
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            )
            """
        )
    conn.commit()


def _ensure_project_deleted_at_schema(conn: Any) -> None:
    """项目软删除：deleted_at 非空表示在回收站。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS deleted_at TIMESTAMP")
    else:
        cur.execute("PRAGMA table_info(projects)")
        cols = [row[1] for row in cur.fetchall()]
        if "deleted_at" not in cols:
            cur.execute("ALTER TABLE projects ADD COLUMN deleted_at TEXT")
    conn.commit()






def _ensure_customer_engagement_schema(conn: Any) -> None:
    """客户联系人、跟进记录；客户分级/状态/标签（与《03-客户管理》MVP 对齐）。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_contacts (
                id SERIAL PRIMARY KEY,
                customer_id INTEGER NOT NULL REFERENCES customers(id) ON DELETE CASCADE,
                name VARCHAR(120) NOT NULL,
                title VARCHAR(120),
                phone VARCHAR(50),
                email VARCHAR(120),
                is_primary SMALLINT NOT NULL DEFAULT 0,
                note TEXT,
                created_at TIMESTAMP NOT NULL,
                updated_at TIMESTAMP NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_follow_ups (
                id SERIAL PRIMARY KEY,
                customer_id INTEGER NOT NULL REFERENCES customers(id) ON DELETE CASCADE,
                followed_by VARCHAR(120) NOT NULL,
                followed_at TIMESTAMP NOT NULL,
                method VARCHAR(20) NOT NULL DEFAULT 'phone',
                content TEXT NOT NULL,
                next_followup_at DATE,
                opportunity_id INTEGER REFERENCES opportunities(id) ON DELETE SET NULL,
                created_at TIMESTAMP NOT NULL
            )
            """
        )
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS status VARCHAR(30) NOT NULL DEFAULT 'potential'")
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS tier VARCHAR(30) NOT NULL DEFAULT 'normal'")
        cur.execute("ALTER TABLE customers ADD COLUMN IF NOT EXISTS tags VARCHAR(500)")
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                title TEXT,
                phone TEXT,
                email TEXT,
                is_primary INTEGER NOT NULL DEFAULT 0,
                note TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id) ON DELETE CASCADE
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_follow_ups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER NOT NULL,
                followed_by TEXT NOT NULL,
                followed_at TEXT NOT NULL,
                method TEXT NOT NULL DEFAULT 'phone',
                content TEXT NOT NULL,
                next_followup_at TEXT,
                opportunity_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id) ON DELETE CASCADE,
                FOREIGN KEY(opportunity_id) REFERENCES opportunities(id) ON DELETE SET NULL
            )
            """
        )
        cur.execute("PRAGMA table_info(customers)")
        cust_cols = [row[1] for row in cur.fetchall()]
        if "status" not in cust_cols:
            cur.execute("ALTER TABLE customers ADD COLUMN status TEXT NOT NULL DEFAULT 'potential'")
        if "tier" not in cust_cols:
            cur.execute("ALTER TABLE customers ADD COLUMN tier TEXT NOT NULL DEFAULT 'normal'")
        if "tags" not in cust_cols:
            cur.execute("ALTER TABLE customers ADD COLUMN tags TEXT")
    idx = [
        "CREATE INDEX IF NOT EXISTS idx_customer_contacts_customer ON customer_contacts(customer_id)",
        "CREATE INDEX IF NOT EXISTS idx_customer_follow_ups_customer_followed ON customer_follow_ups(customer_id, followed_at)",
        "CREATE INDEX IF NOT EXISTS idx_customer_follow_ups_next ON customer_follow_ups(customer_id, next_followup_at)",
        "CREATE INDEX IF NOT EXISTS idx_customers_owner_tier ON customers(owner, tier)",
    ]
    for statement in idx:
        cur.execute(statement)
    conn.commit()


def _ensure_opportunity_pipeline_schema(conn: Any) -> None:
    """商机销售阶段、阶段进入时间、阶段变更日志；合同关联商机 id（《04-商机管理》MVP / OP-022）。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute("ALTER TABLE opportunities ADD COLUMN IF NOT EXISTS stage VARCHAR(50)")
        cur.execute("ALTER TABLE opportunities ADD COLUMN IF NOT EXISTS stage_started_at TIMESTAMP")
        cur.execute("ALTER TABLE opportunities ADD COLUMN IF NOT EXISTS amount_confidence VARCHAR(20)")
        cur.execute("ALTER TABLE opportunities ADD COLUMN IF NOT EXISTS lost_reason VARCHAR(80)")
        cur.execute("ALTER TABLE opportunities ADD COLUMN IF NOT EXISTS lost_reason_note TEXT")
        cur.execute("ALTER TABLE opportunities ADD COLUMN IF NOT EXISTS competitor TEXT")
        cur.execute(
            """
            ALTER TABLE contracts ADD COLUMN IF NOT EXISTS opportunity_id INTEGER
            REFERENCES opportunities(id) ON DELETE SET NULL
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS opportunity_stage_logs (
                id SERIAL PRIMARY KEY,
                opportunity_id INTEGER NOT NULL REFERENCES opportunities(id) ON DELETE CASCADE,
                from_stage VARCHAR(50),
                to_stage VARCHAR(50) NOT NULL,
                note TEXT,
                changed_by VARCHAR(120) NOT NULL,
                changed_at TIMESTAMP NOT NULL
            )
            """
        )
    else:
        cur.execute("PRAGMA table_info(opportunities)")
        opp_cols = {row[1] for row in cur.fetchall()}
        for col, ddl in (
            ("stage", "ALTER TABLE opportunities ADD COLUMN stage TEXT"),
            ("stage_started_at", "ALTER TABLE opportunities ADD COLUMN stage_started_at TEXT"),
            ("amount_confidence", "ALTER TABLE opportunities ADD COLUMN amount_confidence TEXT"),
            ("lost_reason", "ALTER TABLE opportunities ADD COLUMN lost_reason TEXT"),
            ("lost_reason_note", "ALTER TABLE opportunities ADD COLUMN lost_reason_note TEXT"),
            ("competitor", "ALTER TABLE opportunities ADD COLUMN competitor TEXT"),
        ):
            if col not in opp_cols:
                cur.execute(ddl)
                opp_cols.add(col)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS opportunity_stage_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                opportunity_id INTEGER NOT NULL,
                from_stage TEXT,
                to_stage TEXT NOT NULL,
                note TEXT,
                changed_by TEXT NOT NULL,
                changed_at TEXT NOT NULL,
                FOREIGN KEY(opportunity_id) REFERENCES opportunities(id) ON DELETE CASCADE
            )
            """
        )
    _upd_stage_sql = (
        """
        UPDATE opportunities SET stage = CASE status
            WHEN 'new' THEN 'lead'
            WHEN 'tracking' THEN 'validate'
            WHEN 'won' THEN 'won'
            WHEN 'lost' THEN 'lost'
            ELSE 'validate'
        END
        WHERE stage IS NULL OR TRIM(COALESCE(stage::text, '')) = ''
        """
        if uses_postgres()
        else """
        UPDATE opportunities SET stage = CASE status
            WHEN 'new' THEN 'lead'
            WHEN 'tracking' THEN 'validate'
            WHEN 'won' THEN 'won'
            WHEN 'lost' THEN 'lost'
            ELSE 'validate'
        END
        WHERE stage IS NULL OR TRIM(COALESCE(stage, '')) = ''
        """
    )
    cur.execute(_upd_stage_sql)
    cur.execute(
        """
        UPDATE opportunities SET stage_started_at = updated_at
        WHERE stage_started_at IS NULL
        """
    )
    idx_opp = [
        "CREATE INDEX IF NOT EXISTS idx_opportunities_stage_started ON opportunities(stage, stage_started_at)",
        "CREATE INDEX IF NOT EXISTS idx_opportunities_owner_stage ON opportunities(owner, stage)",
    ]
    for statement in idx_opp:
        cur.execute(statement)
    conn.commit()


def _ensure_contracts_extended_schema(conn: Any) -> None:
    """合同扩展：关联商机、负责人、币种、到期日（与《05-合同管理》CT-004/CT-010 等对齐）。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute(
            """
            ALTER TABLE contracts ADD COLUMN IF NOT EXISTS opportunity_id INTEGER
            REFERENCES opportunities(id) ON DELETE SET NULL
            """
        )
        cur.execute("ALTER TABLE contracts ADD COLUMN IF NOT EXISTS owner VARCHAR(120)")
        cur.execute("ALTER TABLE contracts ADD COLUMN IF NOT EXISTS currency VARCHAR(10) NOT NULL DEFAULT 'CNY'")
        cur.execute("ALTER TABLE contracts ADD COLUMN IF NOT EXISTS end_date DATE")
    else:
        cur.execute("PRAGMA table_info(contracts)")
        cols = [row[1] for row in cur.fetchall()]
        if "opportunity_id" not in cols:
            cur.execute("ALTER TABLE contracts ADD COLUMN opportunity_id INTEGER")
        if "owner" not in cols:
            cur.execute("ALTER TABLE contracts ADD COLUMN owner TEXT")
        if "currency" not in cols:
            cur.execute("ALTER TABLE contracts ADD COLUMN currency TEXT NOT NULL DEFAULT 'CNY'")
        if "end_date" not in cols:
            cur.execute("ALTER TABLE contracts ADD COLUMN end_date TEXT")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_contracts_opportunity ON contracts(opportunity_id)")
    conn.commit()


def _ensure_project_phases_schema(conn: Any) -> None:
    """项目交付阶段相关表"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_delivery_phases (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                name VARCHAR(120) NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                status VARCHAR(50) NOT NULL DEFAULT 'pending',
                description TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS phase_template_items (
                id SERIAL PRIMARY KEY,
                template_id INTEGER NOT NULL,
                name VARCHAR(120) NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                default_duration_days INTEGER NOT NULL DEFAULT 0,
                description TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_phase_change_logs (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                summary TEXT NOT NULL,
                actor VARCHAR(120) NOT NULL,
                created_at TIMESTAMP NOT NULL
            )
            """
        )
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_delivery_phases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pending',
                description TEXT,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS phase_template_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                template_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                default_duration_days INTEGER NOT NULL DEFAULT 0,
                description TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS project_phase_change_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                summary TEXT NOT NULL,
                actor TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            )
            """
        )
    conn.commit()


def _ensure_ai_schema(conn: Any) -> None:
    """AI 生成日志与反馈表：P0 以草稿和审计为主，便于后续替换真实模型。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS ai_generation_logs (
                id SERIAL PRIMARY KEY,
                scene_code VARCHAR(80) NOT NULL,
                target_type VARCHAR(50) NOT NULL,
                target_id INTEGER,
                triggered_by VARCHAR(120) NOT NULL,
                prompt_version VARCHAR(80) NOT NULL,
                provider VARCHAR(80) NOT NULL,
                source_snapshot TEXT,
                generated_content TEXT NOT NULL,
                status VARCHAR(30) NOT NULL DEFAULT 'generated',
                accepted BOOLEAN,
                created_at TIMESTAMP NOT NULL,
                updated_at TIMESTAMP NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS ai_feedback (
                id SERIAL PRIMARY KEY,
                generation_id INTEGER NOT NULL REFERENCES ai_generation_logs(id) ON DELETE CASCADE,
                feedback_type VARCHAR(40) NOT NULL,
                feedback_note TEXT,
                created_by VARCHAR(120) NOT NULL,
                created_at TIMESTAMP NOT NULL
            )
            """
        )
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS ai_generation_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                scene_code TEXT NOT NULL,
                target_type TEXT NOT NULL,
                target_id INTEGER,
                triggered_by TEXT NOT NULL,
                prompt_version TEXT NOT NULL,
                provider TEXT NOT NULL,
                source_snapshot TEXT,
                generated_content TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'generated',
                accepted INTEGER,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS ai_feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                generation_id INTEGER NOT NULL,
                feedback_type TEXT NOT NULL,
                feedback_note TEXT,
                created_by TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(generation_id) REFERENCES ai_generation_logs(id) ON DELETE CASCADE
            )
            """
        )
    conn.commit()


def _ensure_query_indexes(conn: Any) -> None:
    """为高频查询补齐索引，避免首页和列表页在数据增大后退化为全表扫描。"""
    cur = conn.cursor()
    statements = [
        "CREATE INDEX IF NOT EXISTS idx_projects_active_updated ON projects(deleted_at, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_projects_active_status ON projects(deleted_at, status)",
        "CREATE INDEX IF NOT EXISTS idx_projects_active_stage ON projects(deleted_at, current_stage)",
        "CREATE INDEX IF NOT EXISTS idx_tasks_project_updated ON tasks(project_id, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_tasks_project_status_end ON tasks(project_id, status, planned_end)",
        "CREATE INDEX IF NOT EXISTS idx_tasks_depends_on ON tasks(depends_on_task_id)",
        "CREATE INDEX IF NOT EXISTS idx_milestones_project_due_status ON milestones(project_id, due_date, status)",
        "CREATE INDEX IF NOT EXISTS idx_risks_project_level_status ON risks(project_id, level, status)",
        "CREATE INDEX IF NOT EXISTS idx_risks_project_due_status ON risks(project_id, due_date, status)",
        "CREATE INDEX IF NOT EXISTS idx_attachments_project_scope ON attachments(project_id, task_id, milestone_id)",
        "CREATE INDEX IF NOT EXISTS idx_contracts_project_updated ON contracts(project_id, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_contracts_customer_updated ON contracts(customer_id, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_opportunities_customer_updated ON opportunities(customer_id, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_opportunities_status_updated ON opportunities(status, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_approvals_status_updated ON approvals(status, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_receivables_status_plan_date ON receivables(status, plan_date)",
        "CREATE INDEX IF NOT EXISTS idx_invoices_contract_updated ON invoices(contract_id, updated_at)",
        "CREATE INDEX IF NOT EXISTS idx_project_members_project_user ON project_members(project_id, user_id)",
        "CREATE INDEX IF NOT EXISTS idx_project_progress_project_created ON project_progress_entries(project_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_project_activity_project_created ON project_activity_logs(project_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_project_stage_logs_project_changed ON project_stage_logs(project_id, changed_at)",
        "CREATE INDEX IF NOT EXISTS idx_project_delivery_phases_project_sort ON project_delivery_phases(project_id, sort_order)",
        "CREATE INDEX IF NOT EXISTS idx_phase_template_items_template_sort ON phase_template_items(template_id, sort_order)",
        "CREATE INDEX IF NOT EXISTS idx_project_phase_change_logs_project ON project_phase_change_logs(project_id, created_at)",
        "CREATE INDEX IF NOT EXISTS idx_ai_generation_scene_target ON ai_generation_logs(scene_code, target_type, target_id)",
        "CREATE INDEX IF NOT EXISTS idx_ai_generation_created ON ai_generation_logs(created_at)",
        "CREATE INDEX IF NOT EXISTS idx_ai_feedback_generation ON ai_feedback(generation_id, created_at)",
    ]
    for statement in statements:
        cur.execute(statement)
    conn.commit()


def get_db():
    if "db" not in g:
        g.db = create_connection()
        _bootstrap_connection_schema(g.db)
    return g.db


def session_is_system_admin() -> bool:
    """是否为系统管理员（roles.code = admin）。以数据库 users + roles 为准，避免仅依赖 session['role'] 与库不一致。"""
    uid = session.get("user_id")
    if not uid:
        return False
    cached = getattr(g, "_session_is_system_admin", None)
    if cached is not None:
        return bool(cached)
    out = "admin" in set(get_user_role_codes(int(uid)))
    setattr(g, "_session_is_system_admin", out)
    return out


def migrate_sqlite_attachments_columns(conn: Any) -> bool:
    """
    旧版 crm.db 可能缺少 task_id / milestone_id / uploaded_by / file_size，
    此时 INSERT ... uploaded_by, file_size 会失败，文件已写入 uploads 但数据库无记录。
    返回 True 表示执行了 ALTER（调用方需 commit）。
    """
    if uses_postgres():
        return False
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(attachments)")
    cols = [row[1] for row in cur.fetchall()]
    changed = False
    for col, ddl in (
        ("task_id", "ALTER TABLE attachments ADD COLUMN task_id INTEGER"),
        ("milestone_id", "ALTER TABLE attachments ADD COLUMN milestone_id INTEGER"),
        ("uploaded_by", "ALTER TABLE attachments ADD COLUMN uploaded_by TEXT"),
        ("file_size", "ALTER TABLE attachments ADD COLUMN file_size INTEGER"),
    ):
        if col not in cols:
            cur.execute(ddl)
            cols.append(col)
            changed = True
    return changed


def ensure_sqlite_attachments_schema() -> None:
    """在请求内补全 SQLite attachments 表列（与 init_db 逻辑一致）。"""
    if uses_postgres():
        return
    conn = get_db()
    if migrate_sqlite_attachments_columns(conn):
        conn.commit()
        app.logger.warning(
            "已自动为 SQLite attachments 表补齐列（旧库曾缺列会导致上传无数据库记录）",
        )


def _ensure_approval_notes_schema(conn: Any) -> None:
    """补齐 approvals.apply_note 列，用于保存发起审批时的说明情况。"""
    cur = conn.cursor()
    if uses_postgres():
        cur.execute("ALTER TABLE approvals ADD COLUMN IF NOT EXISTS apply_note TEXT")
        return
    cur.execute("PRAGMA table_info(approvals)")
    columns = [row[1] for row in cur.fetchall()]
    if "apply_note" not in columns:
        cur.execute("ALTER TABLE approvals ADD COLUMN apply_note TEXT")


def fetch_approval_candidates(role_codes: list[str]) -> list[dict[str, Any]]:
    """按角色编码获取可选审批人，返回去重后的用户列表。"""
    codes = [str(code).strip() for code in role_codes if str(code).strip()]
    if not codes:
        return []
    placeholders = ", ".join("?" for _ in codes)
    rows = fetchall(
        f"""
        SELECT u.id AS user_id,
               u.username,
               u.display_name,
               u.is_active AS user_is_active,
               r.code AS role_code,
               r.name AS role_name,
               r.is_active AS role_is_active,
               r.sort_order AS role_sort_order
        FROM users u
        JOIN user_roles ur ON ur.user_id = u.id
        JOIN roles r ON r.id = ur.role_id
        WHERE r.code IN ({placeholders})
        ORDER BY u.id ASC, r.sort_order ASC, r.name ASC
        """,
        tuple(codes),
    )
    grouped: dict[int, dict[str, Any]] = {}
    for row in rows:
        if not bool(row.get("user_is_active")) or not bool(row.get("role_is_active")):
            continue
        user_id = int(row["user_id"])
        item = grouped.setdefault(
            user_id,
            {
                "id": user_id,
                "username": str(row.get("username") or ""),
                "display_name": str(row.get("display_name") or ""),
                "role_labels": [],
            },
        )
        label = str(row.get("role_name") or row.get("role_code") or "").strip()
        if label and label not in item["role_labels"]:
            item["role_labels"].append(label)
    return sorted(
        grouped.values(),
        key=lambda item: (str(item.get("display_name") or ""), str(item.get("username") or ""), int(item.get("id") or 0)),
    )


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return func(*args, **kwargs)

    return wrapper


def get_session_permissions() -> frozenset[str]:
    """当前登录用户在本次请求下的权限键集合；优先数据库角色授权，否则回退 ROLE_PERMISSIONS。"""
    if not session.get("user_id"):
        return frozenset()
    if session_is_system_admin():
        return frozenset({"*"})
    cached = getattr(g, "_session_permissions", None)
    if cached is not None:
        return cached
    perms: set[str] = set()
    role_rows = get_user_role_rows(int(session["user_id"]))
    role_ids = [int(role_row["id"]) for role_row in role_rows if role_row.get("id") is not None]
    if role_ids:
        placeholders = ", ".join("?" for _ in role_ids)
        rows = fetchall(
            f"SELECT role_id, perm_key FROM role_permissions WHERE role_id IN ({placeholders})",
            tuple(role_ids),
        )
        if rows:
            perms.update(str(r["perm_key"]) for r in rows)
    if not perms:
        role = (session.get("role") or "normal").strip()
        perms = set(ROLE_PERMISSIONS.get(role, set()))
    if "*" in perms:
        return frozenset({"*"})
    g._session_permissions = frozenset(perms)
    return g._session_permissions


def has_module_permission(module: str, action: str = "view") -> bool:
    permissions = get_session_permissions()
    if "*" in permissions:
        return True
    if "all:view" in permissions and action == "view":
        return True
    if f"{module}:{action}" in permissions:
        return True
    if action == "view" and f"{module}:manage" in permissions:
        return True
    return False


def permission_required(module: str, action: str = "view"):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login", next=request.path))
            if not has_module_permission(module, action):
                flash("无权限执行该操作。", "danger")
                return redirect(url_for("dashboard"))
            return func(*args, **kwargs)

        return wrapper

    return decorator


def admin_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        if not session_is_system_admin():
            flash("无权限：仅管理员可执行该操作。", "danger")
            return redirect(url_for("dashboard"))
        return func(*args, **kwargs)

    return wrapper


@app.teardown_appcontext
def close_db(_error: object) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


@app.before_request
def require_login():
    open_endpoints = {"login", "logout", "static"}
    if request.endpoint in open_endpoints:
        return
    if request.endpoint is None:
        return
    if session.get("user_id"):
        last_activity_ts = _session_last_activity_ts()
        now_ts = datetime.utcnow().timestamp()
        if last_activity_ts is not None and now_ts - last_activity_ts > SESSION_IDLE_TIMEOUT_SECONDS:
            session.clear()
            flash("由于 30 分钟无操作，登录已自动失效，请重新登录。", "warning")
            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return jsonify({"error": "session_expired", "redirect": url_for("login")}), 401
            return redirect(url_for("login", next=request.path))
        _touch_session_activity()
    if session.get("user_id") and session.get("role") == "super_admin":
        session["role"] = "admin"
    if not session.get("user_id"):
        return redirect(url_for("login", next=request.path))


def init_db() -> None:
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    conn = create_connection()
    statements = []
    if uses_postgres():
        statements = [
            """
            CREATE TABLE IF NOT EXISTS projects (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL,
                customer_id INTEGER,
                project_type VARCHAR(50) NOT NULL,
                manager VARCHAR(100) NOT NULL,
                status VARCHAR(50) NOT NULL DEFAULT 'in_progress',
                current_stage VARCHAR(50) NOT NULL DEFAULT 'init',
                start_date DATE,
                end_date DATE,
                description TEXT,
                created_at TIMESTAMP NOT NULL,
                updated_at TIMESTAMP NOT NULL,
                deleted_at TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS tasks (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                assignee VARCHAR(100) NOT NULL,
                status VARCHAR(50) NOT NULL DEFAULT 'todo',
                progress INTEGER NOT NULL DEFAULT 0,
                priority VARCHAR(20) NOT NULL DEFAULT 'medium',
                planned_end DATE,
                actual_end DATE,
                blocked_reason TEXT,
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS milestones (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                owner VARCHAR(100) NOT NULL,
                due_date DATE NOT NULL,
                status VARCHAR(50) NOT NULL DEFAULT 'open',
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS risks (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                title TEXT NOT NULL,
                level VARCHAR(20) NOT NULL DEFAULT 'medium',
                status VARCHAR(50) NOT NULL DEFAULT 'open',
                owner VARCHAR(100) NOT NULL,
                due_date DATE,
                mitigation TEXT,
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS attachments (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE,
                milestone_id INTEGER REFERENCES milestones(id) ON DELETE CASCADE,
                category VARCHAR(50) NOT NULL DEFAULT '合同',
                file_name TEXT NOT NULL,
                stored_name TEXT NOT NULL UNIQUE,
                uploaded_at TIMESTAMP NOT NULL,
                uploaded_by VARCHAR(200),
                file_size BIGINT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS customers (
                id SERIAL PRIMARY KEY,
                name VARCHAR(200) NOT NULL UNIQUE,
                owner VARCHAR(100) NOT NULL,
                phone VARCHAR(50),
                email VARCHAR(120),
                industry VARCHAR(120),
                level VARCHAR(20) NOT NULL DEFAULT 'A',
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS opportunities (
                id SERIAL PRIMARY KEY,
                customer_id INTEGER NOT NULL REFERENCES customers(id) ON DELETE CASCADE,
                title VARCHAR(200) NOT NULL,
                amount NUMERIC(14, 2) NOT NULL DEFAULT 0,
                owner VARCHAR(100) NOT NULL,
                status VARCHAR(50) NOT NULL DEFAULT 'new',
                expected_sign_date DATE,
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS contracts (
                id SERIAL PRIMARY KEY,
                customer_id INTEGER NOT NULL REFERENCES customers(id) ON DELETE CASCADE,
                project_id INTEGER REFERENCES projects(id) ON DELETE SET NULL,
                contract_no VARCHAR(120) NOT NULL,
                amount NUMERIC(14, 2) NOT NULL DEFAULT 0,
                sign_date DATE,
                status VARCHAR(50) NOT NULL DEFAULT 'draft',
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS project_stage_logs (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                stage VARCHAR(50) NOT NULL,
                note TEXT,
                changed_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS approvals (
                id SERIAL PRIMARY KEY,
                module_type VARCHAR(50) NOT NULL,
                module_id INTEGER NOT NULL,
                title VARCHAR(220) NOT NULL,
                requested_value VARCHAR(80) NOT NULL,
                applicant VARCHAR(100) NOT NULL,
                approver VARCHAR(100) NOT NULL,
                status VARCHAR(30) NOT NULL DEFAULT 'pending',
                apply_note TEXT,
                comment TEXT,
                created_at TIMESTAMP NOT NULL,
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS receivables (
                id SERIAL PRIMARY KEY,
                contract_id INTEGER NOT NULL REFERENCES contracts(id) ON DELETE CASCADE,
                plan_date DATE NOT NULL,
                plan_amount NUMERIC(14, 2) NOT NULL DEFAULT 0,
                actual_date DATE,
                actual_amount NUMERIC(14, 2) NOT NULL DEFAULT 0,
                status VARCHAR(30) NOT NULL DEFAULT 'planned',
                note TEXT,
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username VARCHAR(80) NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                display_name VARCHAR(120) NOT NULL,
                role VARCHAR(20) NOT NULL DEFAULT 'normal',
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                updated_at TIMESTAMP NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS invoices (
                id SERIAL PRIMARY KEY,
                invoice_no VARCHAR(120) NOT NULL UNIQUE,
                contract_id INTEGER NOT NULL REFERENCES contracts(id) ON DELETE CASCADE,
                receivable_id INTEGER REFERENCES receivables(id) ON DELETE SET NULL,
                amount NUMERIC(14, 2) NOT NULL DEFAULT 0,
                invoice_date DATE NOT NULL,
                invoice_type VARCHAR(50) NOT NULL,
                invoice_code VARCHAR(120) NOT NULL UNIQUE,
                status VARCHAR(50) NOT NULL DEFAULT 'issued',
                created_by VARCHAR(120) NOT NULL,
                updated_at TIMESTAMP NOT NULL
            )
            """,
        ]
    else:
        statements = [
            """
            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                customer_id INTEGER,
                project_type TEXT NOT NULL,
                manager TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'in_progress',
                current_stage TEXT NOT NULL DEFAULT 'init',
                start_date TEXT,
                end_date TEXT,
                description TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                deleted_at TEXT
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                assignee TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'todo',
                progress INTEGER NOT NULL DEFAULT 0,
                priority TEXT NOT NULL DEFAULT 'medium',
                planned_end TEXT,
                actual_end TEXT,
                blocked_reason TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS milestones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                owner TEXT NOT NULL,
                due_date TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'open',
                updated_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS risks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                level TEXT NOT NULL DEFAULT 'medium',
                status TEXT NOT NULL DEFAULT 'open',
                owner TEXT NOT NULL,
                due_date TEXT,
                mitigation TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                task_id INTEGER,
                milestone_id INTEGER,
                category TEXT NOT NULL DEFAULT '合同',
                file_name TEXT NOT NULL,
                stored_name TEXT NOT NULL UNIQUE,
                uploaded_at TEXT NOT NULL,
                uploaded_by TEXT,
                file_size INTEGER,
                FOREIGN KEY(project_id) REFERENCES projects(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                owner TEXT NOT NULL,
                phone TEXT,
                email TEXT,
                industry TEXT,
                level TEXT NOT NULL DEFAULT 'A',
                updated_at TEXT NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS opportunities (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                owner TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'new',
                stage TEXT NOT NULL DEFAULT 'lead',
                stage_started_at TEXT NOT NULL,
                amount_confidence TEXT,
                lost_reason TEXT,
                lost_reason_note TEXT,
                competitor TEXT,
                expected_sign_date TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS contracts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER NOT NULL,
                project_id INTEGER,
                contract_no TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                sign_date TEXT,
                status TEXT NOT NULL DEFAULT 'draft',
                updated_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id),
                FOREIGN KEY(project_id) REFERENCES projects(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS project_stage_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                stage TEXT NOT NULL,
                note TEXT,
                changed_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS approvals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                module_type TEXT NOT NULL,
                module_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                requested_value TEXT NOT NULL,
                applicant TEXT NOT NULL,
                approver TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending',
                apply_note TEXT,
                comment TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS receivables (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                contract_id INTEGER NOT NULL,
                plan_date TEXT NOT NULL,
                plan_amount REAL NOT NULL DEFAULT 0,
                actual_date TEXT,
                actual_amount REAL NOT NULL DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'planned',
                note TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(contract_id) REFERENCES contracts(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                display_name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'normal',
                is_active INTEGER NOT NULL DEFAULT 1,
                updated_at TEXT NOT NULL
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS invoices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                invoice_no TEXT NOT NULL UNIQUE,
                contract_id INTEGER NOT NULL,
                receivable_id INTEGER,
                amount REAL NOT NULL DEFAULT 0,
                invoice_date TEXT NOT NULL,
                invoice_type TEXT NOT NULL,
                invoice_code TEXT NOT NULL UNIQUE,
                status TEXT NOT NULL DEFAULT 'issued',
                created_by TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(contract_id) REFERENCES contracts(id),
                FOREIGN KEY(receivable_id) REFERENCES receivables(id)
            )
            """,
        ]

    cursor = conn.cursor()
    for statement in statements:
        cursor.execute(statement)

    if uses_postgres():
        cursor.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS customer_id INTEGER")
        cursor.execute("ALTER TABLE projects ADD COLUMN IF NOT EXISTS current_stage VARCHAR(50) NOT NULL DEFAULT 'init'")
        cursor.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS role VARCHAR(20) NOT NULL DEFAULT 'normal'")
        cursor.execute("ALTER TABLE attachments ADD COLUMN IF NOT EXISTS task_id INTEGER REFERENCES tasks(id) ON DELETE CASCADE")
        cursor.execute("ALTER TABLE attachments ADD COLUMN IF NOT EXISTS milestone_id INTEGER REFERENCES milestones(id) ON DELETE CASCADE")
        cursor.execute("ALTER TABLE attachments ADD COLUMN IF NOT EXISTS uploaded_by VARCHAR(200)")
        cursor.execute("ALTER TABLE attachments ADD COLUMN IF NOT EXISTS file_size BIGINT")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS project_activity_logs (
                id SERIAL PRIMARY KEY,
                project_id INTEGER NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
                actor VARCHAR(120) NOT NULL,
                action VARCHAR(100) NOT NULL,
                entity_type VARCHAR(50) NOT NULL,
                entity_label TEXT,
                detail TEXT,
                created_at TIMESTAMP NOT NULL
            )
            """
        )
        cursor.execute(
            "ALTER TABLE tasks ADD COLUMN IF NOT EXISTS depends_on_task_id INTEGER REFERENCES tasks(id) ON DELETE SET NULL"
        )
    else:
        cursor.execute("PRAGMA table_info(projects)")
        columns = [row[1] for row in cursor.fetchall()]
        if "customer_id" not in columns:
            cursor.execute("ALTER TABLE projects ADD COLUMN customer_id INTEGER")
        if "current_stage" not in columns:
            cursor.execute("ALTER TABLE projects ADD COLUMN current_stage TEXT NOT NULL DEFAULT 'init'")
        cursor.execute("PRAGMA table_info(users)")
        user_columns = [row[1] for row in cursor.fetchall()]
        if "role" not in user_columns:
            cursor.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'normal'")
        migrate_sqlite_attachments_columns(conn)
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS project_activity_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                actor TEXT NOT NULL,
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_label TEXT,
                detail TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id)
            )
            """
        )
        cursor.execute("PRAGMA table_info(tasks)")
        task_columns = [row[1] for row in cursor.fetchall()]
        if "depends_on_task_id" not in task_columns:
            cursor.execute("ALTER TABLE tasks ADD COLUMN depends_on_task_id INTEGER")

    _ensure_roles_schema(conn)
    _ensure_project_progress_schema(conn)
    _ensure_project_members_schema(conn)
    _ensure_project_deleted_at_schema(conn)
    _ensure_approval_notes_schema(conn)

    _ensure_customer_engagement_schema(conn)
    _ensure_contracts_extended_schema(conn)
    _ensure_project_phases_schema(conn)
    _ensure_ai_schema(conn)
    _ensure_query_indexes(conn)

    if uses_postgres():
        cursor.execute(
            """
            INSERT INTO users(username, password_hash, display_name, role, is_active, updated_at, role_id)
            VALUES (%s, %s, %s, 'admin', TRUE, %s, (SELECT id FROM roles WHERE code = 'admin' LIMIT 1))
            ON CONFLICT (username) DO NOTHING
            """,
            (DEFAULT_ADMIN_USER, generate_password_hash(DEFAULT_ADMIN_PASSWORD), "系统管理员", now_iso()),
        )
        cursor.execute(
            "UPDATE users SET role = 'admin', role_id = (SELECT id FROM roles WHERE code = 'admin' LIMIT 1), updated_at = %s WHERE username = %s",
            (now_iso(), DEFAULT_ADMIN_USER),
        )
    else:
        cursor.execute(
            """
            INSERT OR IGNORE INTO users(username, password_hash, display_name, role, is_active, updated_at, role_id)
            VALUES (?, ?, ?, 'admin', 1, ?, (SELECT id FROM roles WHERE code = 'admin' LIMIT 1))
            """,
            (DEFAULT_ADMIN_USER, generate_password_hash(DEFAULT_ADMIN_PASSWORD), "系统管理员", now_iso()),
        )
        cursor.execute(
            "UPDATE users SET role = 'admin', role_id = (SELECT id FROM roles WHERE code = 'admin' LIMIT 1), updated_at = ? WHERE username = ?",
            (now_iso(), DEFAULT_ADMIN_USER),
        )

    conn.commit()
    _DB_BOOTSTRAPPED_KEYS.add(_db_bootstrap_key())
    conn.close()


def sql_project_active(alias: str = "p") -> str:
    """仅未进回收站的项目（deleted_at 为空）。"""
    return f"({alias}.deleted_at IS NULL)"


def recycle_cutoff_iso() -> str:
    """早于该时间的软删除记录可被自动彻底清理（与 now_iso 格式一致）。"""
    return (datetime.now() - timedelta(days=PROJECT_RECYCLE_DAYS)).replace(microsecond=0).isoformat(sep=" ")


def purge_project_forever(project_id: int) -> None:
    """从磁盘与数据库彻底删除项目（含附件文件行，项目行删除后级联子表）。"""
    attachments = fetchall("SELECT id, stored_name FROM attachments WHERE project_id = ?", (project_id,))
    for attachment in attachments:
        remove_attachment_file(attachment["stored_name"])
        execute("DELETE FROM attachments WHERE id = ?", (attachment["id"],))
    execute("DELETE FROM projects WHERE id = ?", (project_id,))


def purge_expired_projects() -> int:
    """物理删除回收站中超过 PROJECT_RECYCLE_DAYS 的项目。返回清理数量。"""
    cutoff = recycle_cutoff_iso()
    rows = fetchall(
        "SELECT id FROM projects WHERE deleted_at IS NOT NULL AND deleted_at < ?",
        (cutoff,),
    )
    n = 0
    for row in rows:
        purge_project_forever(int(row["id"]))
        n += 1
    if n:
        get_db().commit()
    return n


def enrich_attachment_for_display(a: dict[str, Any]) -> dict[str, Any]:
    out = dict(a)
    stored = out.get("stored_name")
    if stored and out.get("file_size") is None:
        try:
            cached_size = _ATTACHMENT_FILE_SIZE_CACHE.get(str(stored))
            if cached_size is not None:
                out["file_size"] = cached_size
            else:
                p = UPLOAD_DIR / stored
                if p.is_file():
                    file_size = p.stat().st_size
                    out["file_size"] = file_size
                    with _ATTACHMENT_FILE_SIZE_LOCK:
                        _ATTACHMENT_FILE_SIZE_CACHE[str(stored)] = file_size
        except OSError:
            pass
    return out


def attachment_public_dict(a: dict[str, Any]) -> dict[str, Any]:
    e = enrich_attachment_for_display(a)
    fs = e.get("file_size")
    uploader = e.get("uploaded_by")
    uploader_s = (str(uploader).strip() if uploader not in (None, "") else "") or "—"
    return {
        "id": int(e["id"]),
        "file_name": e["file_name"],
        "category": e.get("category") or "其他",
        "uploaded_at": str(e.get("uploaded_at") or ""),
        "uploaded_by": uploader_s,
        "file_size": fs,
        "file_size_label": format_file_size(fs),
    }


def ensure_excel_support() -> bool:
    return Workbook is not None and load_workbook is not None


def fetch_customer_crm_todos_for_owner(owner_name: str) -> list[dict[str, Any]]:
    """
    客户模块待办 TD-C01～TD-C03（与 customers.owner 与登录用户 display_name 对齐）。
    返回字段：todo_code, item_type, item_name, due_at, customer_id
    """
    if not (owner_name or "").strip():
        return []
    owner = owner_name.strip()
    today = date.today().isoformat()
    stale_cutoff = (date.today() - timedelta(days=CUSTOMER_STALE_FOLLOW_DAYS)).isoformat()
    items: list[dict[str, Any]] = []

    td_c01 = fetchall(
        """
        SELECT c.id AS customer_id, c.name AS customer_name, f.next_followup_at AS due_at
        FROM customers c
        INNER JOIN customer_follow_ups f ON f.id = (
            SELECT f2.id FROM customer_follow_ups f2
            WHERE f2.customer_id = c.id
            ORDER BY f2.followed_at DESC
            LIMIT 1
        )
        WHERE c.owner = ?
          AND f.next_followup_at IS NOT NULL AND TRIM(CAST(f.next_followup_at AS TEXT)) != ''
          AND CAST(f.next_followup_at AS TEXT) <= ?
        """,
        (owner, today),
    )
    td_c02_has_follow = fetchall(
        """
        SELECT c.id AS customer_id, c.name AS customer_name, mf.last_follow AS due_at
        FROM customers c
        INNER JOIN (
            SELECT customer_id, MAX(followed_at) AS last_follow
            FROM customer_follow_ups
            GROUP BY customer_id
        ) mf ON mf.customer_id = c.id
        WHERE c.owner = ?
          AND SUBSTR(CAST(mf.last_follow AS TEXT), 1, 10) <= ?
        """,
        (owner, stale_cutoff),
    )
    td_c02_no_follow = fetchall(
        """
        SELECT c.id AS customer_id, c.name AS customer_name, c.updated_at AS due_at
        FROM customers c
        WHERE c.owner = ?
          AND NOT EXISTS (SELECT 1 FROM customer_follow_ups f WHERE f.customer_id = c.id)
          AND SUBSTR(CAST(c.updated_at AS TEXT), 1, 10) <= ?
        """,
        (owner, stale_cutoff),
    )
    td_c02 = list(td_c02_has_follow) + list(td_c02_no_follow)
    c01_ids = {int(r["customer_id"]) for r in td_c01}

    for row in td_c01:
        items.append(
            {
                "todo_code": "TD-C01",
                "item_type": "跟进到期提醒",
                "item_name": f"{row['customer_name']} · 下次跟进已到期",
                "due_at": str(row.get("due_at") or ""),
                "customer_id": int(row["customer_id"]),
            }
        )

    seen_c02: set[int] = set()
    for row in td_c02:
        cid = int(row["customer_id"])
        if cid in c01_ids or cid in seen_c02:
            continue
        seen_c02.add(cid)
        items.append(
            {
                "todo_code": "TD-C02",
                "item_type": "长期未跟进",
                "item_name": f"{row['customer_name']} · 超{CUSTOMER_STALE_FOLLOW_DAYS}天无新跟进",
                "due_at": str(row.get("due_at") or "")[:19],
                "customer_id": cid,
            }
        )

    td_c03 = fetchall(
        """
        SELECT c.id AS customer_id, c.name AS customer_name
        FROM customers c
        WHERE c.owner = ?
          AND c.tier = 'strategic'
          AND NOT EXISTS (
            SELECT 1 FROM customer_contacts cc
            WHERE cc.customer_id = c.id AND cc.is_primary = 1
          )
        """,
        (owner,),
    )
    for row in td_c03:
        items.append(
            {
                "todo_code": "TD-C03",
                "item_type": "重点客户待完善",
                "item_name": f"{row['customer_name']} · 战略客户缺主联系人",
                "due_at": "",
                "customer_id": int(row["customer_id"]),
            }
        )

    return items


def crm_summary() -> dict[str, int]:
    row = fetchone(
        """
        SELECT
            (SELECT COUNT(1) FROM customers) AS customer_count,
            (SELECT COUNT(1) FROM opportunities) AS opportunity_count,
            (SELECT COUNT(1) FROM contracts) AS contract_count,
            (SELECT COUNT(1) FROM approvals WHERE status = 'pending') AS pending_approval_count,
            (SELECT COUNT(1) FROM receivables WHERE status != 'received' AND plan_date < ?) AS overdue_receivable_count
        """,
        (date.today().isoformat(),),
    )
    return {
        "customer_count": int(row["customer_count"]) if row and row.get("customer_count") is not None else 0,
        "opportunity_count": int(row["opportunity_count"]) if row and row.get("opportunity_count") is not None else 0,
        "contract_count": int(row["contract_count"]) if row and row.get("contract_count") is not None else 0,
        "pending_approval_count": int(row["pending_approval_count"]) if row and row.get("pending_approval_count") is not None else 0,
        "overdue_receivable_count": int(row["overdue_receivable_count"]) if row and row.get("overdue_receivable_count") is not None else 0,
    }


def generate_invoice_no() -> str:
    year = datetime.now().strftime("%Y")
    prefix = f"KP{year}"
    row = fetchone("SELECT COUNT(1) AS c FROM invoices WHERE invoice_no LIKE ?", (f"{prefix}%",))
    seq = int(row["c"]) + 1 if row else 1
    return f"{prefix}{seq:08d}"


def fetch_project_metric_maps(project_ids: list[int]) -> tuple[dict[int, int], dict[int, int], dict[int, int]]:
    """
    批量计算项目进度、延期里程碑数、高风险数。

    之前 dashboard / project_manage 会为每个项目各查 3 次，数据一多就会形成明显的 N+1 查询。
    这里改为按 project_id 聚合，收益高且不改变原有业务语义。
    """
    if not project_ids:
        return {}, {}, {}

    in_clause, params = in_clause_params(project_ids)
    today = date.today().isoformat()
    progress_map = {
        int(row["project_id"]): int(round(float(row["avg_progress"])))
        for row in fetchall(
            f"""
            SELECT project_id, AVG(progress) AS avg_progress
            FROM tasks
            WHERE project_id IN {in_clause}
            GROUP BY project_id
            """,
            params,
        )
        if row.get("avg_progress") is not None
    }
    delayed_map = {
        int(row["project_id"]): int(row["delayed_count"])
        for row in fetchall(
            f"""
            SELECT project_id, COUNT(1) AS delayed_count
            FROM milestones
            WHERE project_id IN {in_clause}
              AND status != 'done'
              AND due_date < ?
            GROUP BY project_id
            """,
            params + (today,),
        )
    }
    high_risk_map = {
        int(row["project_id"]): int(row["risk_count"])
        for row in fetchall(
            f"""
            SELECT project_id, COUNT(1) AS risk_count
            FROM risks
            WHERE project_id IN {in_clause}
              AND status != 'closed'
              AND level = 'high'
            GROUP BY project_id
            """,
            params,
        )
    }
    return progress_map, delayed_map, high_risk_map


def compute_project_progress(project_id: int) -> int:
    row = fetchone("SELECT AVG(progress) AS avg_progress FROM tasks WHERE project_id = ?", (project_id,))
    if row is None or row.get("avg_progress") is None:
        return 0
    return int(round(float(row["avg_progress"])))


def delayed_milestone_count(project_id: int) -> int:
    row = fetchone(
        """
        SELECT COUNT(1) AS c
        FROM milestones
        WHERE project_id = ?
          AND status != 'done'
          AND due_date < ?
        """,
        (project_id, date.today().isoformat()),
    )
    return int(row["c"]) if row else 0


def open_high_risk_count(project_id: int) -> int:
    row = fetchone(
        """
        SELECT COUNT(1) AS c
        FROM risks
        WHERE project_id = ?
          AND status != 'closed'
          AND level = 'high'
        """,
        (project_id,),
    )
    return int(row["c"]) if row else 0


def annotate_projects_with_metrics(projects: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """就地补齐项目进度与健康度，供多个列表页复用。"""
    project_ids = [int(project["id"]) for project in projects]
    progress_map, delayed_map, high_risk_map = fetch_project_metric_maps(project_ids)
    for project in projects:
        project_id = int(project["id"])
        project["progress"] = progress_map.get(project_id, 0)
        project["delayed_milestones"] = delayed_map.get(project_id, 0)
        project["open_high_risks"] = high_risk_map.get(project_id, 0)
        if project["delayed_milestones"] > 0 or project["open_high_risks"] > 0:
            project["health"] = "red"
        elif project["progress"] < 60:
            project["health"] = "yellow"
        else:
            project["health"] = "green"
    return projects


def touch_project(project_id: int) -> None:
    execute("UPDATE projects SET updated_at = ? WHERE id = ? AND deleted_at IS NULL", (now_iso(), project_id))


def current_actor_name() -> str:
    return session.get("display_name") or session.get("username") or "系统"


def _serialize_ai_payload(value: Any) -> str:
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, default=str)


def log_ai_generation(
    scene_code: str,
    target_type: str,
    target_id: int | None,
    source_snapshot: Any,
    generated_content: Any,
    *,
    status: str = "generated",
    prompt_version: str = "p0-v1",
    provider: str = "heuristic-local",
    triggered_by: str | None = None,
) -> int:
    actor = (triggered_by or current_actor_name() or "系统")[:120]
    created_at = now_iso()
    cur = get_db().cursor()
    params = (
        (scene_code or "")[:80],
        (target_type or "")[:50],
        target_id,
        actor,
        (prompt_version or "p0-v1")[:80],
        (provider or "heuristic-local")[:80],
        _serialize_ai_payload(source_snapshot),
        _serialize_ai_payload(generated_content),
        (status or "generated")[:30],
        created_at,
        created_at,
    )
    if uses_postgres():
        cur.execute(
            q(
                """
                INSERT INTO ai_generation_logs(
                    scene_code, target_type, target_id, triggered_by, prompt_version, provider,
                    source_snapshot, generated_content, status, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                RETURNING id
                """
            ),
            params,
        )
        row = cur.fetchone()
        if not row:
            raise RuntimeError("AI 生成日志写入失败。")
        return int(as_dict(row)["id"])
    cur.execute(
        q(
            """
            INSERT INTO ai_generation_logs(
                scene_code, target_type, target_id, triggered_by, prompt_version, provider,
                source_snapshot, generated_content, status, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
        ),
        params,
    )
    return int(cur.lastrowid or 0)


def set_ai_generation_accepted(generation_id: int, accepted: bool) -> None:
    execute(
        "UPDATE ai_generation_logs SET accepted = ?, updated_at = ? WHERE id = ?",
        (sql_bool(accepted), now_iso(), generation_id),
    )


def user_is_project_manager(project: dict[str, Any]) -> bool:
    """当前用户是否为该项目经理（管理员也可填写项目进展）。"""
    if session_is_system_admin():
        return True
    mgr = (project.get("manager") or "").strip().lower()
    if not mgr:
        return False
    dn = (session.get("display_name") or "").strip().lower()
    un = (session.get("username") or "").strip().lower()
    return dn == mgr or un == mgr


def can_manage_project_members(project: dict[str, Any]) -> bool:
    """项目成员增删：仅项目经理本人、项目总监或系统管理员。"""
    return can_manage_project_record(project)


def _sql_user_active(alias: str = "u") -> str:
    if uses_postgres():
        return f"COALESCE({alias}.is_active, TRUE) IS TRUE"
    return f"COALESCE({alias}.is_active, 1) = 1"


def _sql_effective_role_on_user(alias_u: str = "u") -> str:
    return f"COALESCE(NULLIF(TRIM(r.code), ''), NULLIF(TRIM({alias_u}.role), ''))"


def search_pm_users(q: str, limit: int = 30) -> list[dict[str, Any]]:
    """用户管理中角色为项目经理（pm）的用户，支持模糊匹配姓名/账号。"""
    act = _sql_user_active("u")
    role_expr = _sql_effective_role_on_user("u")
    pat = f"%{(q or '').strip().lower()}%"
    if not (q or "").strip():
        pat = "%"
    order = f"{sql_user_order_clause('u')}, u.username"
    sql = f"""
        SELECT u.id, u.username, u.display_name
        FROM users u
        LEFT JOIN roles r ON r.id = u.role_id
        WHERE {act}
        AND {role_expr} = ?
        AND (LOWER(COALESCE(u.display_name, '')) LIKE ? OR LOWER(COALESCE(u.username, '')) LIKE ?)
        ORDER BY {order}
        LIMIT ?
    """
    return fetchall(sql, (PM_ROLE_CODE, pat, pat, limit))


def search_active_users(q: str, limit: int = 30) -> list[dict[str, Any]]:
    """所有在职用户的模糊搜索结果，用于通用人员选择下拉。"""
    act = _sql_user_active("u")
    pat = f"%{(q or '').strip().lower()}%"
    if not (q or "").strip():
        pat = "%"
    order = f"{sql_user_order_clause('u')}, u.username"
    sql = f"""
        SELECT u.id, u.username, u.display_name
        FROM users u
        WHERE {act}
        AND (LOWER(COALESCE(u.display_name, '')) LIKE ? OR LOWER(COALESCE(u.username, '')) LIKE ?)
        ORDER BY {order}
        LIMIT ?
    """
    return fetchall(sql, (pat, pat, limit))


def search_member_candidates(project_id: int, q: str, limit: int = 30) -> list[dict[str, Any]]:
    """可加入项目的用户（未在成员表中），模糊匹配。"""
    act = _sql_user_active("u")
    pat = f"%{(q or '').strip().lower()}%"
    if not (q or "").strip():
        pat = "%"
    order = f"{sql_user_order_clause('u')}, u.username"
    sql = f"""
        SELECT u.id, u.username, u.display_name
        FROM users u
        WHERE {act}
        AND u.id NOT IN (SELECT user_id FROM project_members WHERE project_id = ?)
        AND (LOWER(COALESCE(u.display_name, '')) LIKE ? OR LOWER(COALESCE(u.username, '')) LIKE ?)
        ORDER BY {order}
        LIMIT ?
    """
    return fetchall(sql, (project_id, pat, pat, limit))


def user_is_pm_user(user_id: int) -> bool:
    act = _sql_user_active("u")
    role_expr = _sql_effective_role_on_user("u")
    row = fetchone(
        f"""
        SELECT 1 AS ok FROM users u
        LEFT JOIN roles r ON r.id = u.role_id
        WHERE u.id = ?
        AND {act}
        AND {role_expr} = ?
        """,
        (user_id, PM_ROLE_CODE),
    )
    return row is not None


def manager_display_string(user_id: int) -> str:
    u = fetchone("SELECT display_name, username FROM users WHERE id = ?", (user_id,))
    if not u:
        return ""
    return (u.get("display_name") or u.get("username") or "").strip()


def current_user_identity_values() -> frozenset[str]:
    """当前用户可被业务记录引用的身份字符串，统一做小写和空白归一化。"""
    values: set[str] = set()
    for raw in (
        session.get("display_name"),
        session.get("username"),
        session.get("role"),
    ):
        text = str(raw or "").strip().lower()
        if text:
            values.add(text)
    return frozenset(values)


def current_user_matches_text(value: object | None) -> bool:
    """判断文本负责人/责任人字段是否匹配当前用户。"""
    text = str(value or "").strip().lower()
    if not text:
        return False
    return text in current_user_identity_values()


def current_user_role_codes() -> set[str]:
    uid = session.get("user_id")
    if not uid:
        return set()
    return set(get_user_role_codes(int(uid)))


def approval_requested_value_label(value: object | None) -> str:
    """审批请求值的展示标签。"""
    return {
        "won": "商机赢单",
        "signed": "合同签约",
        "close_project": "项目结项",
    }.get(str(value or "").strip(), str(value or "").strip() or "-")


def approval_visible_requested_values(role_codes: set[str] | None = None) -> set[str] | None:
    """
    返回当前用户可见的审批请求值集合。

    返回 None 表示不过滤全部可见；返回空集合表示没有任何审批可见。
    """
    codes = set(role_codes or current_user_role_codes())
    if session_is_system_admin() or codes & {"admin", "management"}:
        return None
    visible: set[str] = set()
    if "sales_director" in codes:
        visible.update({"won", "signed"})
    if "project_director" in codes:
        visible.add("close_project")
    return visible


def _sql_identity_in_clause(column_sql: str, values: set[str]) -> tuple[str, list[str]]:
    cleaned = sorted({str(v).strip().lower() for v in values if str(v).strip()})
    if not cleaned:
        return "0=1", []
    placeholders = ", ".join("?" for _ in cleaned)
    return f"LOWER(TRIM(COALESCE({column_sql}, ''))) IN ({placeholders})", cleaned


def build_customer_visibility_clause(alias: str = "c") -> tuple[str, list[object]]:
    """返回 customer 列表/详情可见范围的 SQL 片段。"""
    if session_is_system_admin():
        return "1=1", []
    role_codes = current_user_role_codes()
    if role_codes & {"management", "sales_director", "project_director", "finance"}:
        return "1=1", []
    if role_codes & {"sales", "pm", "implementer"}:
        return "1=1", []
    return "0=1", []


def build_opportunity_visibility_clause(alias: str = "o") -> tuple[str, list[object]]:
    if session_is_system_admin():
        return "1=1", []
    role_codes = current_user_role_codes()
    if role_codes & {"management", "sales_director", "project_director", "finance"}:
        return "1=1", []
    if role_codes & {"sales", "pm", "implementer"}:
        return "1=1", []
    return "0=1", []


def build_contract_visibility_clause(alias: str = "ct") -> tuple[str, list[object]]:
    if session_is_system_admin():
        return "1=1", []
    role_codes = current_user_role_codes()
    if role_codes & {"management", "sales_director", "project_director", "finance"}:
        return "1=1", []
    if role_codes & {"sales", "pm", "implementer"}:
        return "1=1", []
    return "0=1", []


def build_project_visibility_clause(alias: str = "p") -> tuple[str, list[object]]:
    if session_is_system_admin():
        return "1=1", []
    role_codes = current_user_role_codes()
    if role_codes & {"management", "sales_director", "project_director", "finance"}:
        return "1=1", []
    if "sales" in role_codes:
        return "1=1", []
    clauses: list[str] = []
    params: list[object] = []
    identities = current_user_identity_values()
    if identities:
        clause, values = _sql_identity_in_clause(f"{alias}.manager", identities)
        clauses.append(clause)
        params.extend(values)
    if session.get("user_id"):
        clauses.append(
            f"EXISTS (SELECT 1 FROM project_members pm WHERE pm.project_id = {alias}.id AND pm.user_id = ?)"
        )
        params.append(int(session["user_id"]))
    if clauses:
        return "(" + " OR ".join(clauses) + ")", params
    return "0=1", []


def can_manage_project_record(project: dict[str, Any]) -> bool:
    """是否可管理某个项目的基础信息与交付内容。"""
    if session_is_system_admin():
        return True
    role_codes = current_user_role_codes()
    if "project_director" in role_codes:
        return True
    return user_is_project_manager(project)


def find_pm_user_id_for_manager_label(manager_label: str) -> int | None:
    """将项目表中的经理展示字段匹配到项目经理角色用户 id。"""
    if not (manager_label or "").strip():
        return None
    m = manager_label.strip()
    row = fetchone(
        f"""
        SELECT u.id FROM users u
        LEFT JOIN roles r ON r.id = u.role_id
        WHERE {_sql_effective_role_on_user("u")} = ?
        AND (LOWER(TRIM(u.display_name)) = LOWER(?) OR LOWER(TRIM(u.username)) = LOWER(?))
        """,
        (PM_ROLE_CODE, m, m),
    )
    return int(row["id"]) if row else None


def log_project_activity(
    project_id: int,
    action: str,
    entity_type: str,
    entity_label: str = "",
    detail: str | None = None,
) -> None:
    label = (entity_label or "")[:500]
    detail_text = (detail or "")[:2000] if detail else None
    execute(
        """
        INSERT INTO project_activity_logs(project_id, actor, action, entity_type, entity_label, detail, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (project_id, current_actor_name(), action, entity_type, label, detail_text, now_iso()),
    )


def log_project_phase_change(project_id: int, summary: str) -> None:
    execute(
        """
        INSERT INTO project_phase_change_logs(project_id, summary, actor, created_at)
        VALUES (?, ?, ?, ?)
        """,
        (project_id, (summary or "")[:2000], current_actor_name(), now_iso()),
    )


def _normalize_require_all_deliverables(raw: Any) -> Any:
    b = True if raw is None else (bool(raw) if isinstance(raw, bool) else int(raw) != 0)
    if uses_postgres():
        return b
    return 1 if b else 0




def fetch_project_attention_items(limit_total: int = 12, project_ids: list[int] | set[int] | tuple[int, ...] | None = None) -> list[dict[str, Any]]:
    """聚合多项目下的延期里程碑、逾期风险、逾期任务、项目计划结束日超期。"""
    today = date.today().isoformat()
    items: list[dict[str, Any]] = []
    seen: set[tuple[str, int, str]] = set()
    scoped_ids = sorted({int(pid) for pid in (project_ids or []) if str(pid).strip()})
    scoped_clause = ""
    scoped_params: list[object] = []
    if scoped_ids:
        placeholders = ", ".join("?" for _ in scoped_ids)
        scoped_clause = f"AND p.id IN ({placeholders})"
        scoped_params = list(scoped_ids)

    def push(kind: str, name: str, due: str | None, pid: int, pname: str) -> bool:
        if len(items) >= limit_total:
            return False
        key = (kind, pid, name)
        if key in seen:
            return True
        seen.add(key)
        items.append(
            {
                "item_type": kind,
                "item_name": name,
                "due_at": due,
                "project_id": pid,
                "project_name": pname,
                "link": url_for("project_detail", project_id=pid),
            }
        )
        return True

    for row in fetchall(
        """
        SELECT m.title AS ttitle, m.due_date AS due_at, p.id AS project_id, p.name AS project_name
        FROM milestones m
        JOIN projects p ON p.id = m.project_id
        WHERE p.deleted_at IS NULL
          AND p.status != 'closed'
          {scoped_clause}
          AND m.status IN ('open', 'delayed')
          AND m.due_date < ?
        ORDER BY m.due_date ASC
        LIMIT 30
        """.format(scoped_clause=scoped_clause),
        tuple(scoped_params) + (today,),
    ):
        if not push("逾期里程碑", row["ttitle"], str(row["due_at"]), int(row["project_id"]), row["project_name"]):
            return items

    for row in fetchall(
        """
        SELECT r.title AS ttitle, r.due_date AS due_at, p.id AS project_id, p.name AS project_name
        FROM risks r
        JOIN projects p ON p.id = r.project_id
        WHERE p.deleted_at IS NULL
          {scoped_clause}
          AND r.status IN ('open', 'in_progress')
          AND r.due_date IS NOT NULL
          AND r.due_date < ?
        ORDER BY r.due_date ASC
        LIMIT 30
        """.format(scoped_clause=scoped_clause),
        tuple(scoped_params) + (today,),
    ):
        if not push("逾期风险跟进", row["ttitle"], str(row["due_at"]), int(row["project_id"]), row["project_name"]):
            return items

    for row in fetchall(
        """
        SELECT t.title AS ttitle, t.planned_end AS due_at, p.id AS project_id, p.name AS project_name
        FROM tasks t
        JOIN projects p ON p.id = t.project_id
        WHERE p.deleted_at IS NULL
          AND p.status != 'closed'
          {scoped_clause}
          AND t.status != 'done'
          AND t.planned_end IS NOT NULL
          AND t.planned_end < ?
        ORDER BY t.planned_end ASC
        LIMIT 30
        """.format(scoped_clause=scoped_clause),
        tuple(scoped_params) + (today,),
    ):
        if not push("逾期任务", row["ttitle"], str(row["due_at"]), int(row["project_id"]), row["project_name"]):
            return items

    for row in fetchall(
        """
        SELECT p.name AS ttitle, p.end_date AS due_at, p.id AS project_id, p.name AS project_name
        FROM projects p
        WHERE p.deleted_at IS NULL
          AND p.status != 'closed'
          {scoped_clause}
          AND p.end_date IS NOT NULL
          AND p.end_date < ?
        ORDER BY p.end_date ASC
        LIMIT 30
        """.format(scoped_clause=scoped_clause),
        tuple(scoped_params) + (today,),
    ):
        if not push("项目计划已超期", row["ttitle"], str(row["due_at"]), int(row["project_id"]), row["project_name"]):
            return items

    return items


def parse_task_depends(project_id: int, raw: str | None, exclude_task_id: int | None = None) -> int | None:
    if raw is None or str(raw).strip() == "":
        return None
    try:
        tid = int(str(raw).strip())
    except ValueError:
        return None
    if exclude_task_id is not None and tid == exclude_task_id:
        return None
    row = fetchone("SELECT id, project_id FROM tasks WHERE id = ?", (tid,))
    if not row or int(row["project_id"]) != project_id:
        return None
    return tid


def parse_int_form_value(raw: Any, default: int | None = None) -> int | None:
    """安全解析表单整数值；非法输入返回 default，避免直接抛出 500。"""
    if raw is None:
        return default
    text = str(raw).strip()
    if text == "":
        return default
    try:
        return int(text)
    except (TypeError, ValueError):
        return default


def parse_float_form_value(raw: Any, default: float | None = None) -> float | None:
    """安全解析表单浮点值；非法输入返回 default，避免直接抛出 500。"""
    if raw is None:
        return default
    text = str(raw).strip()
    if text == "":
        return default
    try:
        return float(text)
    except (TypeError, ValueError):
        return default


def instantiate_project_phases_from_template(project_id: int, template_id: int) -> None:
    items = fetchall(
        "SELECT name, sort_order, default_duration_days, description FROM phase_template_items WHERE template_id = ?",
        (template_id,),
    )
    if not items:
        raise ValueError("模板无阶段定义")
    execute("DELETE FROM project_delivery_phases WHERE project_id = ? AND status = 'pending'", (project_id,))
    for item in items:
        execute(
            """
            INSERT INTO project_delivery_phases(project_id, name, sort_order, status, description)
            VALUES (?, ?, ?, 'pending', ?)
            """,
            (project_id, item["name"], item["sort_order"], item["description"]),
        )


def has_pending_approval(module_type: str, module_id: int, requested_value: str) -> bool:
    row = fetchone(
        """
        SELECT id
        FROM approvals
        WHERE module_type = ? AND module_id = ? AND requested_value = ? AND status = 'pending'
        LIMIT 1
        """,
        (module_type, module_id, requested_value),
    )
    return row is not None


def submit_approval(
    module_type: str,
    module_id: int,
    title: str,
    requested_value: str,
    applicant: str,
    approver: str,
    apply_note: str | None = None,
) -> None:
    execute(
        """
        INSERT INTO approvals(module_type, module_id, title, requested_value, applicant, approver, status, apply_note, comment, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, 'pending', ?, NULL, ?, ?)
        """,
        (module_type, module_id, title, requested_value, applicant, approver, apply_note, now_iso(), now_iso()),
    )


def insert_opportunity_stage_log(
    opportunity_id: int,
    from_stage: str | None,
    to_stage: str,
    note: str | None,
    changed_by: str,
) -> None:
    execute(
        """
        INSERT INTO opportunity_stage_logs(opportunity_id, from_stage, to_stage, note, changed_by, changed_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (opportunity_id, from_stage, to_stage, note, changed_by, now_iso()),
    )


def apply_approval(approval: dict[str, Any]) -> None:
    module_type = approval["module_type"]
    module_id = int(approval["module_id"])
    requested_value = approval["requested_value"]
    if module_type == "opportunity":
        row = fetchone("SELECT stage FROM opportunities WHERE id = ?", (module_id,))
        from_stage = str(row["stage"]) if row and row.get("stage") is not None else None
        ts = now_iso()
        execute(
            """
            UPDATE opportunities SET status = ?, stage = 'won', stage_started_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (requested_value, ts, ts, module_id),
        )
        if requested_value == "won":
            actor = session.get("display_name") or session.get("username") or str(approval.get("approver") or "")
            insert_opportunity_stage_log(module_id, from_stage, "won", "审批通过", actor)
    elif module_type == "contract":
        execute(
            "UPDATE contracts SET status = ?, updated_at = ? WHERE id = ?",
            (requested_value, now_iso(), module_id),
        )
    elif module_type == "project":
        if requested_value == "close_project":
            execute(
                "UPDATE projects SET status = 'closed', current_stage = 'closed', updated_at = ? WHERE id = ? AND deleted_at IS NULL",
                (now_iso(), module_id),
            )
            execute(
                "INSERT INTO project_stage_logs(project_id, stage, note, changed_at) VALUES (?, 'closed', ?, ?)",
                (module_id, "审批通过，项目结项", now_iso()),
            )


def remove_attachment_file(stored_name: str) -> None:
    path = UPLOAD_DIR / stored_name
    if path.exists():
        path.unlink(missing_ok=True)



# 共享 helper 全部定义完成后再加载路由模块，避免循环导入时取到半初始化状态。
import routes_business  # noqa: F401
import routes_projects  # noqa: F401
import routes_system  # noqa: F401

if __name__ == "__main__":
    if not DB_PATH.exists():
        os.makedirs(BASE_DIR, exist_ok=True)
    init_db()
    app.run(host=APP_HOST, port=APP_PORT, debug=True)
